import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def export_ledger_xlsx(data: dict, output_path: str):
    """
    data: output dari LedgerFormatter.format_journal_entries()
    output_path: path file .xlsx hasil export
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "General Ledger"

    # ── Styles ──────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", size=10, bold=True)
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # ── Title block ─────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = "SMALL BUSINESS GENERAL LEDGER TEMPLATE"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "ACCOUNT NAME"
    ws["B3"] = data.get("account_name", "")
    ws["A4"] = "ACCT NO."
    ws["B4"] = data.get("acct_no", "")
    ws["E3"] = "MONTH ENDING"
    ws["F3"] = data.get("month_ending", "")
    ws["E4"] = "STARTING BALANCE"
    ws["F4"] = data.get("starting_balance", 0)
    ws["E5"] = "TOTAL ADJUSTED BALANCE"
    ws["F5"] = data.get("total_adjusted", 0)

    for row in range(3, 6):
        for col in ["A", "B", "E", "F"]:
            cell = ws[f"{col}{row}"]
            cell.font = Font(name="Calibri", size=10, bold=True)

    # ── Table header ────────────────────────────────────────────────────
    headers = ["DATE", "DESCRIPTION", "POST REFERENCE", "TRANSACTIONS", "", "BALANCES", ""]
    sub_headers = ["", "", "", "DEBIT", "CREDIT", "TOTAL DEBIT", "TOTAL CREDIT"]

    for col_idx, (h, sh) in enumerate(zip(headers, sub_headers), start=1):
        cell_top = ws.cell(row=7, column=col_idx, value=h)
        cell_top.font = header_font
        cell_top.fill = header_fill
        cell_top.alignment = center_align
        cell_top.border = thin_border

        cell_sub = ws.cell(row=8, column=col_idx, value=sh if sh else None)
        cell_sub.font = header_font
        cell_sub.fill = header_fill
        cell_sub.alignment = center_align
        cell_sub.border = thin_border

    # Merge header cells
    ws.merge_cells("D7:E7")  # TRANSACTIONS spans DEBIT & CREDIT
    ws.merge_cells("F7:G7")  # BALANCES spans TOTAL DEBIT & TOTAL CREDIT

    # ── Data rows ───────────────────────────────────────────────────────
    start_row = 9
    for i, row in enumerate(data.get("rows", [])):
        r = start_row + i
        ws.cell(row=r, column=1, value=row.get("date", ""))
        ws.cell(row=r, column=2, value=row.get("description", ""))
        ws.cell(row=r, column=3, value=row.get("post_ref", ""))
        ws.cell(row=r, column=4, value=row.get("debit", 0))
        ws.cell(row=r, column=5, value=row.get("credit", 0))
        ws.cell(row=r, column=6, value=row.get("running_debit", 0))
        ws.cell(row=r, column=7, value=row.get("running_credit", 0))
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = thin_border
            ws.cell(row=r, column=col).alignment = center_align

    # ── Column widths ───────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"Exported to {output_path}")
